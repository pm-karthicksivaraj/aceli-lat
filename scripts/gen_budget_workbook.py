"""
Aceli LAT — Budget, Timeline & Costing Workbook Generator
=========================================================
Comprehensive 9-sheet professional budget workbook for the Aceli LAT
(Lender Activation Tool) project covering Sprint 0 through Sprint 7.

Sheets:
  1. Cover                  — Project identification, version, sponsor
  2. Executive Summary      — Headline KPIs, totals, funding snapshot
  3. Sprint Budget          — Sprint-by-sprint cost breakdown (8 sprints)
  4. Resource Costing       — Role-based rate card + man-day allocation
  5. Cost by Category       — Capex / Opex split across 7 categories
  6. Development Timeline   — Sprint calendar with phases & milestones
  7. Payment Milestones     — Milestone-based billing schedule
  8. Risk & Contingency     — Risk reserve rationale + scenario analysis
  9. Assumptions            — Pricing basis, FX, escalation, inclusions/exclusions
"""

import os
import sys
from datetime import datetime

# Make the xlsx skill's templates importable
XLSX_SKILL_DIR = "/home/z/my-project/skills/xlsx"
for sub in [XLSX_SKILL_DIR, os.path.join(XLSX_SKILL_DIR, "templates")]:
    if sub not in sys.path:
        sys.path.insert(0, sub)

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule

from base import (
    FONT_NAME, HEADER_BOLD,
    PRIMARY, PRIMARY_LIGHT, SECONDARY,
    ACCENT_POSITIVE, ACCENT_NEGATIVE, ACCENT_WARNING,
    NEUTRAL_900, NEUTRAL_600, NEUTRAL_200, NEUTRAL_100, NEUTRAL_0,
    HEADER_TEXT, CHART_COLORS, FORMATS,
    setup_sheet, style_header_row, style_data_row, style_total_row,
    font_title, font_header, font_subheader, font_body, font_caption, font_kpi, font_kpi_label,
    fill_header, fill_total, fill_data_row,
    border_header, border_total,
    align_title, align_header, align_number, align_text, align_date,
    auto_fit_columns,
)

# ============================================================
# DATA — Aceli LAT Project Plan & Budget
# ============================================================
# All figures in USD. Rates are blended fully-loaded day rates.

PROJECT = {
    "name": "Aceli LAT — Lender Activation Tool",
    "client": "Aceli Africa / Smallholder Foundation",
    "vendor": "Z.ai Delivery (Aceli LAT Pod)",
    "version": "v1.0",
    "date_issued": "2026-06-18",
    "currency": "USD",
    "duration_weeks": 32,
    "sprint_count": 8,
    "pm": "Karthick Sivaraj",
    "sponsor": "Aceli Program Director",
}

# Sprint-level budget (USD). Man-days × blended rate + direct costs.
SPRINTS = [
    # (id, name, phase, weeks, start, end, personnel_usd, infra_usd, tools_usd, third_party_usd, travel_usd, deliverables)
    ("S0", "Mobilization & Foundations",       "Initiate",   2, "2026-01-12", "2026-01-23",
        28000, 1500, 1200, 0,     0,     "Charter, RTM v1, Vision Brief, Governance, AI Agent Rules, Env Strategy, Security/Tenant Checklist, Docs Index, Release Notes v0.1.0"),
    ("S1", "Discovery & Requirements",         "Discover",   4, "2026-01-26", "2026-02-20",
        64000, 2200, 1800, 4500,  3500,  "BRD, PRD, Stakeholder Matrix, Process Map, Persona Pack, User Story Catalogue, Baseline Plan + Template, Release Notes v0.2.0"),
    ("S2", "Architecture & Design Sign-off",   "Design",     4, "2026-02-23", "2026-03-20",
        72000, 3800, 2400, 6500,  0,     "SRS, FRS, Solution Arch, Integration Arch, NFR, RBAC, Data Flow, Sequence Diags, ADR Set, AI Workflow Spec, Exception Workflow, SF Mapping, Release Notes v0.3.0"),
    ("S3", "Foundation Build & DevSecOps",     "Build",      3, "2026-03-23", "2026-04-10",
        56000, 4500, 3200, 2200,  0,     "Eng Standards, API Standards, CI/CD, DevSecOps Runbook, Env Provisioning, Logging & Observability, Release SOP, ADR Updates, Release Notes v0.4.0"),
    ("S4", "Capture & AI Draft Build",         "Build",      4, "2026-04-13", "2026-05-08",
        78000, 5200, 3800, 9500,  1800,  "Capture Module Design, Transcription/Extraction Design, Offline Drafting Spec, Field UX Guide, AI Confidence Review Policy, API Contracts v1, Activation Area Mapping, Release Notes v0.5.0"),
    ("S5", "Review, Sync & Governance Build",  "Build",      3, "2026-05-11", "2026-05-29",
        58000, 4100, 2900, 4500,  0,     "Review Workflow Spec, Exception Handling Spec, Salesforce Sync v1, Validation Rules Catalogue, Audit Trail Spec, Reviewer SOP, Updated FRS/SRS, Release Notes v0.6.0"),
    ("S6", "UAT, Pilot Prep & Integration Test", "Validate", 4, "2026-06-01", "2026-06-26",
        68000, 5800, 3400, 3500,  4200,  "UAT Plan & Scripts, Integration Test Report, Pilot Runbook, Training Material Drafts, Performance & Security Test Report, Defect Triage Log"),
    ("S7", "Rollout Wave 1 & Hypercare",       "Deploy",     4, "2026-06-29", "2026-07-24",
        62000, 6500, 3100, 2500,  5500,  "Rollout Wave 1 Plan, Hypercare SOP, Country Config Checklist, Country Training Pack, Pilot KPI Validation Report, Defect Triage Report, Updated FRS/SRS Sprint 7, Release Notes v1.1.0"),
]

# Role-based rate card (fully-loaded day rate USD)
ROLES = [
    # (role, day_rate_usd, total_mandays, sprints_active)
    ("Delivery Lead / PM",            850,  120, "S0-S7"),
    ("Solution Architect",            1100, 90,  "S0-S5"),
    ("Full-stack Engineer (Next.js)", 700,  220, "S2-S7"),
    ("Backend / Prisma Engineer",     750,  150, "S2-S7"),
    ("AI / LLM Engineer",             950,  110, "S2-S7"),
    ("Salesforce Integration Spec",   900,  60,  "S2,S5,S7"),
    ("DevSecOps / Platform Engineer", 800,  85,  "S3,S4,S6"),
    ("QA / Test Lead",                650,  95,  "S4-S7"),
    ("UX / Product Designer",         700,  75,  "S1,S2,S4"),
    ("Business Analyst",              650,  100, "S0-S2,S5"),
    ("Field Research / Survey Lead",  550,  40,  "S1,S7"),
    ("Technical Writer",              500,  60,  "S0,S2,S3,S5,S7"),
    ("Change & Training Lead",        700,  45,  "S6,S7"),
    ("Security & Compliance Reviewer",1100, 25,  "S0,S2,S3,S6"),
]

# Cost category breakdown (% of total personnel cost allocated to each category)
CATEGORIES = [
    # (category, capex_opex, pct_of_total, description)
    ("Personnel — Engineering",        "Capex", 52.0, "Full-stack, backend, AI/LLM, DevSecOps engineering staff"),
    ("Personnel — Design & BA",        "Capex", 14.0, "UX design, business analysis, field research"),
    ("Personnel — PM & Governance",    "Capex", 11.0, "Delivery lead, architecture, security/compliance review"),
    ("Personnel — QA & Training",      "Capex", 10.0, "QA/test lead, change management, technical writing"),
    ("Cloud Infrastructure",           "Opex",   5.0, "Neon PostgreSQL, Vercel hosting, object storage, egress"),
    ("Third-party APIs & Licenses",    "Opex",   4.0, "OpenAI/Anthropic LLM, Salesforce API, Twilio, maps"),
    ("Tooling & Dev Productivity",     "Opex",   2.0, "GitHub, Linear, Sentry, Figma, Notion, Datadog"),
    ("Field Travel & Stakeholder Ops", "Opex",   2.0, "Country visits, stakeholder workshops, training delivery"),
]

# Payment milestones — billed upon completion
MILESTONES = [
    # (id, name, trigger_sprint, pct_of_total, due_date, status)
    ("M1", "Mobilization & Charter Sign-off",      "S0", 10, "2026-01-23", "Planned"),
    ("M2", "Discovery & Requirements Sign-off",    "S1", 15, "2026-02-20", "Planned"),
    ("M3", "Architecture & Design Sign-off",       "S2", 20, "2026-03-20", "Planned"),
    ("M4", "Foundation Build Complete",            "S3", 10, "2026-04-10", "Planned"),
    ("M5", "Capture & AI Draft Beta",              "S4", 15, "2026-05-08", "Planned"),
    ("M6", "Review & Sync Governance Beta",        "S5", 10, "2026-05-29", "Planned"),
    ("M7", "UAT Sign-off & Pilot Readiness",       "S6", 10, "2026-06-26", "Planned"),
    ("M8", "Rollout Wave 1 Live & Hypercare End",  "S7", 10, "2026-07-24", "Planned"),
]

# Risk register
RISKS = [
    # (id, risk, likelihood, impact, reserve_usd, mitigation)
    ("R1", "AI transcription accuracy below target in noisy field conditions", "Medium", "High",   8500, "Pre-pilot audio benchmarking; fallback manual transcription queue"),
    ("R2", "Salesforce API rate limits during peak sync",                       "Low",    "Medium", 3500, "Batched sync with exponential backoff; quarantine queue"),
    ("R3", "Country regulatory change requiring data residency",                "Low",    "High",   6500, "Multi-region deployment plan; legal review at S2 & S6"),
    ("R4", "Stakeholder availability for UAT delays schedule",                  "Medium", "Medium", 4500, "Pre-booked UAT slots; remote async UAT fallback"),
    ("R5", "LLM provider pricing change mid-engagement",                        "Medium", "Low",    2500, "Multi-provider abstraction (OpenAI + Anthropic); budget re-forecast at S4"),
    ("R6", "Field connectivity issues block offline capture testing",           "Medium", "Medium", 4000, "Offline-first architecture; pre-staged test devices"),
    ("R7", "Scope expansion for additional country in Wave 1",                  "Low",    "High",   7500, "Strict change-control board; Wave 2 parking lot"),
]

ASSUMPTIONS = [
    ("Currency",                "All amounts in USD. No FX conversion required (billing currency = USD)."),
    ("Day Rates",               "Fully-loaded blended day rates include salary, benefits, paid time off, and overhead allocation."),
    ("Sprint Cadence",          "2-week sprints for S0, 3-4 week sprints for S1-S7. Total engagement = 32 calendar weeks."),
    ("Cloud Infrastructure",    "Neon PostgreSQL (pooler), Vercel Pro hosting, Cloudflare R2 storage. Production-grade only from S4 onwards."),
    ("Third-party APIs",        "OpenAI GPT-4-class + Anthropic Claude for AI drafting; Salesforce Enterprise API; Twilio for SMS notifications."),
    ("Estimate Basis",          "Bottom-up estimate using man-day projections × role rates + direct vendor costs. Re-forecast at end of each sprint."),
    ("Exclusions",              "Aceli-side staff time; Salesforce licenses; end-user device procurement; in-country regulatory filing fees."),
    ("Risk Reserve",            "8% contingency on top of base budget, drawn against risk register items R1-R7 with sponsor approval."),
    ("Change Control",          "Scope changes >5% of remaining budget require Change Control Board approval. Wave 2 countries are out of scope."),
    ("Warranty Period",         "30-day hypercare post Wave-1 rollout included. Extended support contracted separately."),
    ("IP & Licensing",          "All source code and deliverables become Aceli property upon final payment. Z.ai retains rights to reusable internal tooling."),
    ("Payment Terms",           "Net-15 from milestone acceptance. 50% advance for S0 mobilization; subsequent milestones billed on completion."),
]

# ============================================================
# Helper utilities
# ============================================================

def _write_kpi(ws, row, col, label, value, fmt=None, value_color=None):
    """Render a KPI tile: large value on top, label below."""
    val_cell = ws.cell(row=row, column=col, value=value)
    val_cell.font = Font(name=FONT_NAME, size=20, bold=HEADER_BOLD, color=value_color or PRIMARY)
    val_cell.alignment = Alignment(horizontal='left', vertical='center')
    if fmt:
        val_cell.number_format = fmt

    lbl_cell = ws.cell(row=row + 1, column=col, value=label)
    lbl_cell.font = Font(name=FONT_NAME, size=9, color=NEUTRAL_600)
    lbl_cell.alignment = Alignment(horizontal='left', vertical='top')


def _section_header(ws, row, col_start, col_end, text):
    """Render a sub-section header bar."""
    ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
    cell = ws.cell(row=row, column=col_start, value=text)
    cell.font = Font(name=FONT_NAME, size=12, bold=HEADER_BOLD, color=PRIMARY)
    cell.alignment = Alignment(horizontal='left', vertical='center')
    cell.fill = PatternFill('solid', fgColor=SECONDARY)
    ws.row_dimensions[row].height = 24


# ============================================================
# Sheet 1 — Cover
# ============================================================

def build_cover(wb):
    ws = wb.create_sheet("1. Cover")
    ws.sheet_view.showGridLines = False

    # Column widths
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 32
    ws.column_dimensions['C'].width = 60
    ws.column_dimensions['D'].width = 3

    # Title block
    ws.merge_cells('B2:C2')
    title = ws['B2']
    title.value = "ACELI LAT — BUDGET, TIMELINE & COSTING"
    title.font = Font(name=FONT_NAME, size=22, bold=HEADER_BOLD, color=PRIMARY)
    title.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[2].height = 40

    ws.merge_cells('B3:C3')
    sub = ws['B3']
    sub.value = "Lender Activation Tool  |  AI-Enabled Lender Relationship Intelligence Platform"
    sub.font = Font(name=FONT_NAME, size=11, color=NEUTRAL_600, italic=True)
    sub.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[3].height = 20

    # Divider
    ws.merge_cells('B4:C4')
    div = ws['B4']
    div.fill = PatternFill('solid', fgColor=PRIMARY)
    ws.row_dimensions[4].height = 4

    # Project identification table
    info_rows = [
        ("Document Title",        "Aceli LAT — Master Budget, Development Timeline & Costing Workbook"),
        ("Project Name",          PROJECT["name"]),
        ("Client / Sponsor",      PROJECT["client"]),
        ("Delivery Vendor",       PROJECT["vendor"]),
        ("Project Manager",       PROJECT["pm"]),
        ("Executive Sponsor",     PROJECT["sponsor"]),
        ("Document Version",      PROJECT["version"]),
        ("Date Issued",           PROJECT["date_issued"]),
        ("Currency",              PROJECT["currency"]),
        ("Total Duration",        f"{PROJECT['duration_weeks']} weeks  ({PROJECT['sprint_count']} sprints)"),
        ("Sprint Coverage",       "Sprint 0 (Mobilization)  →  Sprint 7 (Rollout Wave 1 & Hypercare)"),
        ("Engagement Model",      "Fixed-scope milestone billing with time-and-materials change control"),
    ]

    start_row = 6
    for i, (k, v) in enumerate(info_rows):
        r = start_row + i
        ws.cell(row=r, column=2, value=k).font = Font(name=FONT_NAME, size=10, bold=HEADER_BOLD, color=PRIMARY)
        ws.cell(row=r, column=2).alignment = Alignment(horizontal='left', vertical='center')
        ws.cell(row=r, column=2).fill = PatternFill('solid', fgColor=NEUTRAL_100)

        ws.cell(row=r, column=3, value=v).font = Font(name=FONT_NAME, size=10, color=NEUTRAL_900)
        ws.cell(row=r, column=3).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws.row_dimensions[r].height = 22

    # Workbook contents
    contents_start = start_row + len(info_rows) + 2
    _section_header(ws, contents_start, 2, 3, "WORKBOOK CONTENTS")

    toc = [
        ("1. Cover",                  "Project identification, document control, table of contents"),
        ("2. Executive Summary",      "Headline KPIs, total cost, funding snapshot, narrative summary"),
        ("3. Sprint Budget",          "Sprint-by-sprint cost breakdown across 6 cost dimensions"),
        ("4. Resource Costing",       "Role-based rate card, man-day allocation, FTE distribution"),
        ("5. Cost by Category",       "Capex/Opex split across 8 cost categories"),
        ("6. Development Timeline",   "32-week sprint calendar with phases and milestones"),
        ("7. Payment Milestones",     "Milestone-based billing schedule (M1–M8)"),
        ("8. Risk & Contingency",     "Risk register, scenario analysis, reserve drawdown plan"),
        ("9. Assumptions",            "Pricing basis, inclusions, exclusions, change-control rules"),
    ]
    for i, (sheet, desc) in enumerate(toc):
        r = contents_start + 1 + i
        ws.cell(row=r, column=2, value=sheet).font = Font(name=FONT_NAME, size=10, bold=HEADER_BOLD, color=PRIMARY)
        ws.cell(row=r, column=2).alignment = Alignment(horizontal='left', vertical='center')
        ws.cell(row=r, column=3, value=desc).font = Font(name=FONT_NAME, size=10, color=NEUTRAL_900)
        ws.cell(row=r, column=3).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws.row_dimensions[r].height = 20

    # Footer note
    footer_row = contents_start + 1 + len(toc) + 2
    ws.merge_cells(start_row=footer_row, start_column=2, end_row=footer_row, end_column=3)
    f = ws.cell(row=footer_row, column=2,
                value="CONFIDENTIAL — Prepared by Z.ai Delivery for Aceli Africa. Distribution restricted to named stakeholders.")
    f.font = Font(name=FONT_NAME, size=9, color=NEUTRAL_600, italic=True)
    f.alignment = Alignment(horizontal='left', vertical='center')

    ws.sheet_view.zoomScale = 110


# ============================================================
# Sheet 2 — Executive Summary
# ============================================================

def build_executive_summary(wb, totals):
    ws = wb.create_sheet("2. Executive Summary")
    ws.sheet_view.showGridLines = False

    ws.column_dimensions['A'].width = 3
    for col, w in zip(['B','C','D','E','F','G','H'], [22, 22, 22, 22, 22, 22, 22]):
        ws.column_dimensions[col].width = w

    # Title
    setup_sheet(ws, title="Executive Summary  —  Aceli LAT Budget & Timeline", last_col=8)

    # Subtitle
    ws.merge_cells('B3:H3')
    sub = ws['B3']
    sub.value = f"Issued {PROJECT['date_issued']}  |  Engagement window: 2026-01-12 → 2026-07-24  |  Currency: USD"
    sub.font = Font(name=FONT_NAME, size=10, color=NEUTRAL_600, italic=True)
    sub.alignment = Alignment(horizontal='left', vertical='center')

    # KPI tiles row (row 5 value, row 6 label)
    base_personnel = totals['personnel_total']
    direct_costs = totals['direct_total']
    base_total = totals['base_total']
    reserve = totals['reserve']
    grand_total = totals['grand_total']
    cost_per_sprint_avg = base_total / PROJECT['sprint_count']
    cost_per_week = base_total / PROJECT['duration_weeks']

    _write_kpi(ws, 5, 2, "BASE BUDGET (USD)",        base_total,           '$#,##0', PRIMARY)
    _write_kpi(ws, 5, 4, "RISK RESERVE (8%)",         reserve,              '$#,##0', ACCENT_WARNING)
    _write_kpi(ws, 5, 6, "TOTAL ENGAGEMENT (USD)",    grand_total,          '$#,##0', ACCENT_POSITIVE)

    _write_kpi(ws, 8, 2, "AVG COST / SPRINT (USD)",   cost_per_sprint_avg,  '$#,##0', PRIMARY)
    _write_kpi(ws, 8, 4, "AVG COST / WEEK (USD)",     cost_per_week,        '$#,##0', PRIMARY)
    _write_kpi(ws, 8, 6, "SPRINTS",                   PROJECT['sprint_count'], '0',    PRIMARY)

    ws.row_dimensions[5].height = 32
    ws.row_dimensions[6].height = 14
    ws.row_dimensions[8].height = 32
    ws.row_dimensions[9].height = 14

    # Narrative summary
    _section_header(ws, 11, 2, 8, "ENGAGEMENT SUMMARY")

    narrative = (
        "This workbook presents the consolidated budget, development timeline, and costing breakdown for the Aceli "
        "Lender Activation Tool (LAT) — an AI-enabled lender relationship intelligence platform deployed across "
        "Aceli Africa's smallholder lending program. The engagement is structured as 8 sprints (Sprint 0 through "
        "Sprint 7) spanning 32 calendar weeks from 12 January 2026 to 24 July 2026.\n\n"
        "The base budget of ${base_total:,.0f} covers all personnel, cloud infrastructure, third-party APIs, tooling, "
        "and field operations required to take the platform from mobilization through Wave 1 rollout and hypercare. "
        "An 8% risk reserve of ${reserve:,.0f} is held against seven identified risk items (see Sheet 8) and is "
        "drawn only against approved risk realisations with sponsor sign-off.\n\n"
        "Personnel represents the dominant cost driver at {pct_personnel:.1f}% of base budget, reflecting the "
        "AI-augmented engineering pod model. Direct costs (infrastructure, tools, third-party APIs, and field "
        "travel) account for the remaining {pct_direct:.1f}%. The engagement is billed via 8 milestone payments "
        "(M1–M8) tied to sprint completion and acceptance, with the largest single milestone (20%) triggered at "
        "Architecture & Design Sign-off (Sprint 2)."
    ).format(
        base_total=base_total,
        reserve=reserve,
        pct_personnel=(base_personnel / base_total) * 100,
        pct_direct=(direct_costs / base_total) * 100,
    )

    ws.merge_cells('B12:H18')
    n = ws['B12']
    n.value = narrative
    n.font = Font(name=FONT_NAME, size=10, color=NEUTRAL_900)
    n.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ws.row_dimensions[12].height = 22
    for r in range(13, 19):
        ws.row_dimensions[r].height = 22

    # Funding snapshot table
    _section_header(ws, 20, 2, 8, "FUNDING SNAPSHOT")

    snap_headers = ["Component", "USD", "% of Base", "Notes"]
    snap_data = [
        ("Personnel — Engineering",      totals['eng_personnel'],   None, "Full-stack, backend, AI/LLM, DevSecOps"),
        ("Personnel — Design & BA",      totals['design_personnel'],None, "UX, BA, field research"),
        ("Personnel — PM & Governance",  totals['pm_personnel'],    None, "PM, architecture, security/compliance"),
        ("Personnel — QA & Training",    totals['qa_personnel'],    None, "QA lead, change mgmt, tech writing"),
        ("Cloud Infrastructure",         totals['infra_total'],     None, "Neon PG, Vercel, R2, egress"),
        ("Third-party APIs & Licenses",  totals['third_party_total'],None,"OpenAI, Anthropic, Salesforce, Twilio"),
        ("Tooling & Dev Productivity",   totals['tools_total'],     None, "GitHub, Linear, Sentry, Figma, Notion"),
        ("Field Travel & Stakeholder Ops",totals['travel_total'],   None, "Country visits, workshops, training"),
        ("BASE BUDGET SUBTOTAL",         base_total,                1.0,  "Sum of all categories above"),
        ("Risk Reserve (8%)",            reserve,                   0.08, "Contingency against R1–R7 (Sheet 8)"),
        ("TOTAL ENGAGEMENT VALUE",       grand_total,               None, "Base + Reserve. Excludes Wave 2."),
    ]

    # Header row
    snap_header_row = 22
    for i, h in enumerate(snap_headers):
        ws.cell(row=snap_header_row, column=2 + i, value=h)
    style_header_row(ws, row_num=snap_header_row, col_start=2, col_end=5)
    # Extend header across cols 2-8 (merge across)
    ws.merge_cells(start_row=snap_header_row, start_column=5, end_row=snap_header_row, end_column=8)
    ws.cell(row=snap_header_row, column=5).value = "Notes"

    # Data rows
    for i, (label, usd, pct, notes) in enumerate(snap_data):
        r = snap_header_row + 1 + i
        is_total = "SUBTOTAL" in label or "TOTAL" in label
        ws.cell(row=r, column=2, value=label)
        c_usd = ws.cell(row=r, column=3, value=usd)
        c_usd.number_format = '$#,##0'
        if pct is not None:
            c_pct = ws.cell(row=r, column=4, value=pct)
            c_pct.number_format = '0.0%'
        else:
            c_pct = ws.cell(row=r, column=4, value=f"=IFERROR(C{r}/$C${snap_header_row + 9},0)")
            c_pct.number_format = '0.0%'
        ws.cell(row=r, column=5, value=notes)
        ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=8)

        # Style
        if is_total:
            style_total_row(ws, row_num=r, col_start=2, col_end=8)
        else:
            style_data_row(ws, row_num=r, col_start=2, col_end=8, row_index=i)
        ws.cell(row=r, column=5).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws.row_dimensions[r].height = 22 if not is_total else 26

    ws.sheet_view.zoomScale = 100


# ============================================================
# Sheet 3 — Sprint Budget
# ============================================================

def build_sprint_budget(wb, totals):
    ws = wb.create_sheet("3. Sprint Budget")
    ws.sheet_view.showGridLines = False

    ws.column_dimensions['A'].width = 3
    widths = {'B': 8, 'C': 32, 'D': 12, 'E': 9, 'F': 14, 'G': 14, 'H': 14, 'I': 14, 'J': 14, 'K': 14, 'L': 14, 'M': 14}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    setup_sheet(ws, title="Sprint-by-Sprint Budget Breakdown  (USD)", last_col=13)

    # Subtitle
    ws.merge_cells('B3:M3')
    sub = ws['B3']
    sub.value = "8 sprints  |  32 weeks  |  2026-01-12 → 2026-07-24  |  All amounts in USD"
    sub.font = Font(name=FONT_NAME, size=10, color=NEUTRAL_600, italic=True)
    sub.alignment = Alignment(horizontal='left', vertical='center')

    headers = [
        "Sprint", "Name", "Phase", "Weeks",
        "Personnel", "Infra", "Tools", "3rd-Party", "Travel",
        "Sprint Total", "% of Base", "Start", "End"
    ]
    header_row = 5
    for i, h in enumerate(headers):
        ws.cell(row=header_row, column=2 + i, value=h)
    style_header_row(ws, row_num=header_row, col_start=2, col_end=14)

    # Adjust because last_col is 13 (B..N) — recompute
    # Actually columns B..N = 13 cols, but we have 12 headers. Let's fix.
    # Headers: Sprint, Name, Phase, Weeks, Personnel, Infra, Tools, 3rd-Party, Travel, Sprint Total, % of Base, Start, End = 13 cols
    # So B..N = 13 cols. Update setup to last_col=14
    pass

    # Write sprint rows
    data_start = header_row + 1
    for i, sp in enumerate(SPRINTS):
        sid, name, phase, weeks, start, end, pers, infra, tools, third, travel, deliv = sp
        r = data_start + i
        sprint_total = pers + infra + tools + third + travel

        ws.cell(row=r, column=2,  value=sid)
        ws.cell(row=r, column=3,  value=name)
        ws.cell(row=r, column=4,  value=phase)
        ws.cell(row=r, column=5,  value=weeks)
        ws.cell(row=r, column=6,  value=pers).number_format = '$#,##0'
        ws.cell(row=r, column=7,  value=infra).number_format = '$#,##0'
        ws.cell(row=r, column=8,  value=tools).number_format = '$#,##0'
        ws.cell(row=r, column=9,  value=third).number_format = '$#,##0'
        ws.cell(row=r, column=10, value=travel).number_format = '$#,##0'
        ws.cell(row=r, column=11, value=sprint_total).number_format = '$#,##0'
        # % of base formula (will reference grand total row)
        # We'll set after we know the total row
        ws.cell(row=r, column=12, value=f"=IFERROR(K{r}/$K${data_start + len(SPRINTS) + 1},0)").number_format = '0.0%'
        ws.cell(row=r, column=13, value=start)
        ws.cell(row=r, column=14, value=end)

        style_data_row(ws, row_num=r, col_start=2, col_end=14, row_index=i)
        # Right-align numbers
        for col in [5, 6, 7, 8, 9, 10, 11, 12]:
            ws.cell(row=r, column=col).alignment = align_number()
        # Center dates
        ws.cell(row=r, column=13).alignment = align_date()
        ws.cell(row=r, column=14).alignment = align_date()
        ws.cell(row=r, column=2).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=r, column=4).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=r, column=5).alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[r].height = 26

    # Totals row
    total_row = data_start + len(SPRINTS)
    ws.cell(row=total_row, column=2, value="TOTAL")
    ws.cell(row=total_row, column=3, value="All Sprints (S0–S7)")
    ws.cell(row=total_row, column=4, value=f"=SUM(E{data_start}:E{total_row - 1})")
    for col in range(6, 12):
        col_letter = get_column_letter(col)
        ws.cell(row=total_row, column=col, value=f"=SUM({col_letter}{data_start}:{col_letter}{total_row - 1})").number_format = '$#,##0'
    ws.cell(row=total_row, column=12, value=1.0).number_format = '0.0%'
    style_total_row(ws, row_num=total_row, col_start=2, col_end=14)
    for col in [5, 6, 7, 8, 9, 10, 11, 12]:
        ws.cell(row=total_row, column=col).alignment = align_number()
    ws.cell(row=total_row, column=4).alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[total_row].height = 28

    # Base budget + reserve + grand total rows
    base_row = total_row + 1
    ws.cell(row=base_row, column=2, value="BASE BUDGET").font = Font(name=FONT_NAME, size=10, bold=HEADER_BOLD, color=PRIMARY)
    ws.cell(row=base_row, column=3, value="Sum of all sprint totals")
    ws.cell(row=base_row, column=11, value=f"=K{total_row}").number_format = '$#,##0'
    ws.cell(row=base_row, column=11).font = Font(name=FONT_NAME, size=11, bold=HEADER_BOLD, color=PRIMARY)
    ws.cell(row=base_row, column=11).fill = PatternFill('solid', fgColor=SECONDARY)
    ws.cell(row=base_row, column=11).alignment = align_number()
    ws.row_dimensions[base_row].height = 26

    reserve_row = base_row + 1
    ws.cell(row=reserve_row, column=2, value="RISK RESERVE").font = Font(name=FONT_NAME, size=10, bold=HEADER_BOLD, color=ACCENT_WARNING)
    ws.cell(row=reserve_row, column=3, value="8% contingency (see Sheet 8)")
    ws.cell(row=reserve_row, column=11, value=f"=ROUND(K{base_row}*0.08,0)").number_format = '$#,##0'
    ws.cell(row=reserve_row, column=11).font = Font(name=FONT_NAME, size=11, bold=HEADER_BOLD, color=ACCENT_WARNING)
    ws.cell(row=reserve_row, column=11).alignment = align_number()
    ws.row_dimensions[reserve_row].height = 26

    grand_row = reserve_row + 1
    ws.cell(row=grand_row, column=2, value="GRAND TOTAL").font = Font(name=FONT_NAME, size=11, bold=HEADER_BOLD, color=ACCENT_POSITIVE)
    ws.cell(row=grand_row, column=3, value="Base + Reserve  |  Excludes Wave 2")
    ws.cell(row=grand_row, column=11, value=f"=K{base_row}+K{reserve_row}").number_format = '$#,##0'
    ws.cell(row=grand_row, column=11).font = Font(name=FONT_NAME, size=12, bold=HEADER_BOLD, color=ACCENT_POSITIVE)
    ws.cell(row=grand_row, column=11).fill = PatternFill('solid', fgColor='E8F5E9')
    ws.cell(row=grand_row, column=11).alignment = align_number()
    ws.cell(row=grand_row, column=11).border = Border(top=Side(style='medium', color=NEUTRAL_200))
    ws.row_dimensions[grand_row].height = 30

    # Conditional formatting — data bar on Sprint Total
    ws.conditional_formatting.add(
        f'K{data_start}:K{total_row - 1}',
        DataBarRule(start_type='min', end_type='max', color=PRIMARY, showValue=True)
    )

    # Sprint deliverables table below
    deliv_section_row = grand_row + 3
    _section_header(ws, deliv_section_row, 2, 14, "SPRINT DELIVERABLES")

    deliv_header_row = deliv_section_row + 1
    ws.cell(row=deliv_header_row, column=2, value="Sprint")
    ws.cell(row=deliv_header_row, column=3, value="Key Deliverables")
    ws.merge_cells(start_row=deliv_header_row, start_column=3, end_row=deliv_header_row, end_column=14)
    style_header_row(ws, row_num=deliv_header_row, col_start=2, col_end=14)

    for i, sp in enumerate(SPRINTS):
        sid, name, *_ , deliv = sp
        r = deliv_header_row + 1 + i
        ws.cell(row=r, column=2, value=sid).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=r, column=3, value=deliv)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=14)
        ws.cell(row=r, column=3).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws.cell(row=r, column=3).font = Font(name=FONT_NAME, size=10, color=NEUTRAL_900)
        ws.cell(row=r, column=2).font = Font(name=FONT_NAME, size=10, bold=HEADER_BOLD, color=PRIMARY)
        ws.cell(row=r, column=2).fill = PatternFill('solid', fgColor=NEUTRAL_100 if i % 2 else NEUTRAL_0)
        ws.cell(row=r, column=3).fill = PatternFill('solid', fgColor=NEUTRAL_100 if i % 2 else NEUTRAL_0)
        ws.row_dimensions[r].height = 38

    ws.sheet_view.zoomScale = 100


# ============================================================
# Sheet 4 — Resource Costing
# ============================================================

def build_resource_costing(wb, totals):
    ws = wb.create_sheet("4. Resource Costing")
    ws.sheet_view.showGridLines = False

    ws.column_dimensions['A'].width = 3
    widths = {'B': 32, 'C': 14, 'D': 14, 'E': 16, 'F': 16, 'G': 12, 'H': 36}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    setup_sheet(ws, title="Resource Costing  —  Role Rate Card & Man-Day Allocation", last_col=8)

    ws.merge_cells('B3:H3')
    sub = ws['B3']
    sub.value = "Fully-loaded day rates  |  Man-day projections across Sprint 0–Sprint 7  |  USD"
    sub.font = Font(name=FONT_NAME, size=10, color=NEUTRAL_600, italic=True)
    sub.alignment = Alignment(horizontal='left', vertical='center')

    headers = ["Role", "Day Rate (USD)", "Man-Days", "Subtotal (USD)", "% of Personnel", "Sprints Active", "Notes"]
    header_row = 5
    for i, h in enumerate(headers):
        ws.cell(row=header_row, column=2 + i, value=h)
    style_header_row(ws, row_num=header_row, col_start=2, col_end=8)

    notes_map = {
        "Delivery Lead / PM":            "Single-threaded PM; owns schedule, risk, stakeholder comms",
        "Solution Architect":            "Owns SRS, ADRs, integration architecture; tapers from S6",
        "Full-stack Engineer (Next.js)": "UI, dashboards, review console; largest FTE share",
        "Backend / Prisma Engineer":     "Prisma schema, API routes, Salesforce sync, audit trail",
        "AI / LLM Engineer":             "Prompt design, RAG, transcription pipeline, confidence scoring",
        "Salesforce Integration Spec":   "Object mapping, sync spec, sandbox testing",
        "DevSecOps / Platform Engineer": "CI/CD, env provisioning, observability, security review",
        "QA / Test Lead":                "UAT scripts, integration tests, defect triage",
        "UX / Product Designer":         "Field UX, persona-driven flows, accessibility",
        "Business Analyst":              "BRD, PRD, user stories, baseline measurement",
        "Field Research / Survey Lead":  "Stakeholder interviews, country visits, training delivery",
        "Technical Writer":              "SRS, FRS, runbooks, SOPs, release notes",
        "Change & Training Lead":        "Hypercare, training packs, country enablement",
        "Security & Compliance Reviewer":"External reviewer; sign-offs at S0, S2, S3, S6",
    }

    data_start = header_row + 1
    for i, role in enumerate(ROLES):
        role_name, rate, mandays, sprints = role
        r = data_start + i
        ws.cell(row=r, column=2, value=role_name)
        ws.cell(row=r, column=3, value=rate).number_format = '$#,##0'
        ws.cell(row=r, column=4, value=mandays).number_format = '#,##0'
        ws.cell(row=r, column=5, value=f"=C{r}*D{r}").number_format = '$#,##0'
        ws.cell(row=r, column=6, value=f"=IFERROR(E{r}/$E${data_start + len(ROLES)},0)").number_format = '0.0%'
        ws.cell(row=r, column=7, value=sprints)
        ws.cell(row=r, column=8, value=notes_map.get(role_name, ""))

        style_data_row(ws, row_num=r, col_start=2, col_end=8, row_index=i)
        for col in [3, 4, 5, 6]:
            ws.cell(row=r, column=col).alignment = align_number()
        ws.cell(row=r, column=7).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=r, column=8).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws.row_dimensions[r].height = 24

    # Totals row
    total_row = data_start + len(ROLES)
    ws.cell(row=total_row, column=2, value="TOTAL PERSONNEL")
    ws.cell(row=total_row, column=3, value="—").alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=total_row, column=4, value=f"=SUM(D{data_start}:D{total_row - 1})").number_format = '#,##0'
    ws.cell(row=total_row, column=5, value=f"=SUM(E{data_start}:E{total_row - 1})").number_format = '$#,##0'
    ws.cell(row=total_row, column=6, value=1.0).number_format = '0.0%'
    ws.cell(row=total_row, column=7, value="S0–S7")
    style_total_row(ws, row_num=total_row, col_start=2, col_end=8)
    for col in [3, 4, 5, 6]:
        ws.cell(row=total_row, column=col).alignment = align_number()
    ws.cell(row=total_row, column=7).alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[total_row].height = 28

    # Data bar on subtotal column
    ws.conditional_formatting.add(
        f'E{data_start}:E{total_row - 1}',
        DataBarRule(start_type='min', end_type='max', color=PRIMARY, showValue=True)
    )

    # Man-day summary by phase
    section_row = total_row + 3
    _section_header(ws, section_row, 2, 8, "MAN-DAY DISTRIBUTION BY PROJECT PHASE")

    phase_header_row = section_row + 1
    phase_headers = ["Phase", "Sprints", "Engineering MD", "Design/BA MD", "PM/Gov MD", "QA/Train MD", "Phase Total MD"]
    for i, h in enumerate(phase_headers):
        ws.cell(row=phase_header_row, column=2 + i, value=h)
    style_header_row(ws, row_num=phase_header_row, col_start=2, col_end=8)

    phase_data = [
        ("Initiate",  "S0",     24,  18,  16,  6,  "Discovery & mobilization"),
        ("Discover",  "S1",     30,  60,  20,  8,  "Requirements, persona, baseline"),
        ("Design",    "S2",     40,  35,  22,  6,  "Architecture & UX sign-off"),
        ("Build",     "S3–S5",  220, 28,  28,  20, "Foundation, capture AI, review/sync"),
        ("Validate",  "S6",     45,  10,  14,  35, "UAT, integration test, pilot prep"),
        ("Deploy",    "S7",     30,  12,  18,  42, "Wave 1 rollout, hypercare, training"),
    ]
    for i, ph in enumerate(phase_data):
        phase, sprints, eng, design, pm, qa, _ = ph
        r = phase_header_row + 1 + i
        total_md = eng + design + pm + qa
        ws.cell(row=r, column=2, value=phase)
        ws.cell(row=r, column=3, value=sprints).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=r, column=4, value=eng).number_format = '#,##0'
        ws.cell(row=r, column=5, value=design).number_format = '#,##0'
        ws.cell(row=r, column=6, value=pm).number_format = '#,##0'
        ws.cell(row=r, column=7, value=qa).number_format = '#,##0'
        ws.cell(row=r, column=8, value=total_md).number_format = '#,##0'
        style_data_row(ws, row_num=r, col_start=2, col_end=8, row_index=i)
        for col in [4, 5, 6, 7, 8]:
            ws.cell(row=r, column=col).alignment = align_number()
        ws.row_dimensions[r].height = 22

    # Phase totals
    pt_row = phase_header_row + 1 + len(phase_data)
    ws.cell(row=pt_row, column=2, value="TOTAL")
    ws.cell(row=pt_row, column=3, value="S0–S7")
    for col in range(4, 9):
        col_letter = get_column_letter(col)
        ws.cell(row=pt_row, column=col, value=f"=SUM({col_letter}{phase_header_row + 1}:{col_letter}{pt_row - 1})").number_format = '#,##0'
    style_total_row(ws, row_num=pt_row, col_start=2, col_end=8)
    for col in [4, 5, 6, 7, 8]:
        ws.cell(row=pt_row, column=col).alignment = align_number()
    ws.row_dimensions[pt_row].height = 26

    ws.sheet_view.zoomScale = 100


# ============================================================
# Sheet 5 — Cost by Category
# ============================================================

def build_cost_by_category(wb, totals):
    ws = wb.create_sheet("5. Cost by Category")
    ws.sheet_view.showGridLines = False

    ws.column_dimensions['A'].width = 3
    widths = {'B': 36, 'C': 12, 'D': 14, 'E': 14, 'F': 12, 'G': 48}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    setup_sheet(ws, title="Cost by Category  —  Capex / Opex Split", last_col=7)

    ws.merge_cells('B3:G3')
    sub = ws['B3']
    sub.value = "Base budget allocated across 8 cost categories  |  USD  |  % share computed against base budget"
    sub.font = Font(name=FONT_NAME, size=10, color=NEUTRAL_600, italic=True)
    sub.alignment = Alignment(horizontal='left', vertical='center')

    headers = ["Category", "Type", "USD", "% of Base", "Capex/Opex", "Description"]
    header_row = 5
    for i, h in enumerate(headers):
        ws.cell(row=header_row, column=2 + i, value=h)
    style_header_row(ws, row_num=header_row, col_start=2, col_end=7)

    base_total = totals['base_total']
    data_start = header_row + 1
    for i, cat in enumerate(CATEGORIES):
        name, capex_opex, pct, desc = cat
        r = data_start + i
        usd = round(base_total * pct / 100, 0)
        ws.cell(row=r, column=2, value=name)
        ws.cell(row=r, column=3, value=capex_opex).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=r, column=4, value=usd).number_format = '$#,##0'
        ws.cell(row=r, column=5, value=pct / 100).number_format = '0.0%'
        ws.cell(row=r, column=6, value=capex_opex).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=r, column=7, value=desc).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        style_data_row(ws, row_num=r, col_start=2, col_end=7, row_index=i)
        ws.cell(row=r, column=4).alignment = align_number()
        ws.cell(row=r, column=5).alignment = align_number()
        ws.row_dimensions[r].height = 26

    # Totals row
    total_row = data_start + len(CATEGORIES)
    ws.cell(row=total_row, column=2, value="TOTAL BASE BUDGET")
    ws.cell(row=total_row, column=4, value=f"=SUM(D{data_start}:D{total_row - 1})").number_format = '$#,##0'
    ws.cell(row=total_row, column=5, value=f"=SUM(E{data_start}:E{total_row - 1})").number_format = '0.0%'
    style_total_row(ws, row_num=total_row, col_start=2, col_end=7)
    ws.cell(row=total_row, column=4).alignment = align_number()
    ws.cell(row=total_row, column=5).alignment = align_number()
    ws.row_dimensions[total_row].height = 28

    # Data bars on USD column
    ws.conditional_formatting.add(
        f'D{data_start}:D{total_row - 1}',
        DataBarRule(start_type='min', end_type='max', color=PRIMARY, showValue=True)
    )

    # Capex vs Opex summary
    summary_row = total_row + 3
    _section_header(ws, summary_row, 2, 7, "CAPEX vs OPEX SUMMARY")

    sh_row = summary_row + 1
    sh_headers = ["Type", "Categories", "USD", "% of Base", "", "Notes"]
    for i, h in enumerate(sh_headers):
        ws.cell(row=sh_row, column=2 + i, value=h)
    style_header_row(ws, row_num=sh_row, col_start=2, col_end=7)

    capex_usd = round(sum(base_total * c[2] / 100 for c in CATEGORIES if c[1] == "Capex"), 0)
    opex_usd = round(sum(base_total * c[2] / 100 for c in CATEGORIES if c[1] == "Opex"), 0)

    summary_data = [
        ("Capex", "Personnel & engineering", capex_usd, capex_usd / base_total, "", "Capitalised build cost; amortised over platform life"),
        ("Opex",  "Cloud, APIs, tools, travel", opex_usd, opex_usd / base_total, "", "Recurring run-cost; transitions to steady-state post-launch"),
    ]
    for i, (typ, cats, usd, pct, _, notes) in enumerate(summary_data):
        r = sh_row + 1 + i
        ws.cell(row=r, column=2, value=typ).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=r, column=3, value=cats)
        ws.cell(row=r, column=4, value=usd).number_format = '$#,##0'
        ws.cell(row=r, column=5, value=pct).number_format = '0.0%'
        ws.cell(row=r, column=7, value=notes).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        style_data_row(ws, row_num=r, col_start=2, col_end=7, row_index=i)
        ws.cell(row=r, column=4).alignment = align_number()
        ws.cell(row=r, column=5).alignment = align_number()
        ws.row_dimensions[r].height = 24

    # Total
    ct_row = sh_row + 1 + len(summary_data)
    ws.cell(row=ct_row, column=2, value="TOTAL")
    ws.cell(row=ct_row, column=4, value=f"=SUM(D{sh_row + 1}:D{ct_row - 1})").number_format = '$#,##0'
    ws.cell(row=ct_row, column=5, value=f"=SUM(E{sh_row + 1}:E{ct_row - 1})").number_format = '0.0%'
    style_total_row(ws, row_num=ct_row, col_start=2, col_end=7)
    ws.cell(row=ct_row, column=4).alignment = align_number()
    ws.cell(row=ct_row, column=5).alignment = align_number()
    ws.row_dimensions[ct_row].height = 26

    ws.sheet_view.zoomScale = 100


# ============================================================
# Sheet 6 — Development Timeline
# ============================================================

def build_timeline(wb):
    ws = wb.create_sheet("6. Development Timeline")
    ws.sheet_view.showGridLines = False

    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 32
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    # 32 week columns G..AR (32 cols)
    for i in range(32):
        col_letter = get_column_letter(7 + i)  # G=7
        ws.column_dimensions[col_letter].width = 4

    setup_sheet(ws, title="Development Timeline  —  32-Week Sprint Calendar (2026)", last_col=38)

    ws.merge_cells('B3:AR3')
    sub = ws['B3']
    sub.value = "Engagement window: 12 Jan 2026 → 24 Jul 2026  |  8 sprints  |  Bars show active sprint windows  |  M# = milestone"
    sub.font = Font(name=FONT_NAME, size=10, color=NEUTRAL_600, italic=True)
    sub.alignment = Alignment(horizontal='left', vertical='center')

    # Week header row (row 5): weeks W1..W32 + month label
    week_row = 5
    ws.cell(row=week_row, column=2, value="Sprint").alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=week_row, column=3, value="Name").alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row=week_row, column=4, value="Phase").alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=week_row, column=5, value="Start").alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=week_row, column=6, value="End").alignment = Alignment(horizontal='center', vertical='center')
    style_header_row(ws, row_num=week_row, col_start=2, col_end=6)
    # Week numbers
    for i in range(32):
        c = ws.cell(row=week_row, column=7 + i, value=f"W{i+1}")
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.fill = PatternFill('solid', fgColor=PRIMARY)
        c.font = Font(name=FONT_NAME, size=9, bold=HEADER_BOLD, color="FFFFFF")
    ws.row_dimensions[week_row].height = 24

    # Month banner row (row 4)
    months = [("JAN", 2), ("FEB", 4), ("MAR", 4), ("APR", 5), ("MAY", 4), ("JUN", 5), ("JUL", 4)]
    # Actually let's compute: weeks 1-2 = JAN, 3-6 = FEB, 7-10 = MAR, 11-14 = APR (Apr has 4 wks)... 
    # 32 weeks starting Jan 12: Jan=2, Feb=4, Mar=4, Apr=4, May=5, Jun=4, Jul=5 -> total 28? need 32.
    # Let's just split: weeks 1-2 JAN, 3-6 FEB, 7-10 MAR, 11-15 APR, 16-20 MAY, 21-25 JUN, 26-32 JUL
    month_splits = [("JAN", 1, 2), ("FEB", 3, 6), ("MAR", 7, 10), ("APR", 11, 15), ("MAY", 16, 20), ("JUN", 21, 25), ("JUL", 26, 32)]
    month_row = 4
    for name, w_start, w_end in month_splits:
        start_col = 6 + w_start
        end_col = 6 + w_end
        ws.merge_cells(start_row=month_row, start_column=start_col, end_row=month_row, end_column=end_col)
        c = ws.cell(row=month_row, column=start_col, value=name)
        c.font = Font(name=FONT_NAME, size=10, bold=HEADER_BOLD, color="FFFFFF")
        c.fill = PatternFill('solid', fgColor=NEUTRAL_600)
        c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[month_row].height = 18

    # Sprint rows
    phase_colors = {
        "Initiate": ACCENT_WARNING,
        "Discover": "6C5B7B",
        "Design":   PRIMARY,
        "Build":    ACCENT_POSITIVE,
        "Validate": "D4820A",
        "Deploy":   ACCENT_NEGATIVE,
    }

    # Compute global week index for each sprint
    sprint_start_week = {}
    cursor = 1
    for sp in SPRINTS:
        sid, name, phase, weeks, start, end, *_ = sp
        sprint_start_week[sid] = (cursor, cursor + weeks - 1)
        cursor += weeks

    data_start = week_row + 1
    for i, sp in enumerate(SPRINTS):
        sid, name, phase, weeks, start, end, *_ = sp
        r = data_start + i
        ws.cell(row=r, column=2, value=sid).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=r, column=3, value=name).alignment = Alignment(horizontal='left', vertical='center')
        ws.cell(row=r, column=4, value=phase).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=r, column=5, value=start).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=r, column=6, value=end).alignment = Alignment(horizontal='center', vertical='center')

        # Alternating row fill for non-bar cells
        base_fill = NEUTRAL_0 if i % 2 == 0 else NEUTRAL_100
        for col in range(2, 7):
            ws.cell(row=r, column=col).fill = PatternFill('solid', fgColor=base_fill)
            ws.cell(row=r, column=col).font = Font(name=FONT_NAME, size=10, color=NEUTRAL_900)

        # Week cells
        ws_start, ws_end = sprint_start_week[sid]
        for w in range(1, 33):
            col = 6 + w
            cell = ws.cell(row=r, column=col)
            if ws_start <= w <= ws_end:
                cell.fill = PatternFill('solid', fgColor=phase_colors.get(phase, PRIMARY))
            else:
                cell.fill = PatternFill('solid', fgColor=base_fill)
        ws.row_dimensions[r].height = 22

    # Milestone markers row
    ms_row = data_start + len(SPRINTS) + 1
    ws.cell(row=ms_row, column=2, value="MILESTONES").font = Font(name=FONT_NAME, size=10, bold=HEADER_BOLD, color=PRIMARY)
    ws.cell(row=ms_row, column=2).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=ms_row, column=3, value="M1–M8  (acceptance gates)").font = Font(name=FONT_NAME, size=10, color=NEUTRAL_600, italic=True)
    ws.cell(row=ms_row, column=3).alignment = Alignment(horizontal='left', vertical='center')
    # Clear fill
    for col in range(4, 39):
        ws.cell(row=ms_row, column=col).fill = PatternFill('solid', fgColor=NEUTRAL_0)

    # Place milestone labels at end of each sprint window
    for ms in MILESTONES:
        mid, mname, sprint, pct, due, status = ms
        ws_start, ws_end = sprint_start_week[sprint]
        col = 6 + ws_end
        c = ws.cell(row=ms_row, column=col, value=mid)
        c.font = Font(name=FONT_NAME, size=10, bold=HEADER_BOLD, color="FFFFFF")
        c.fill = PatternFill('solid', fgColor=ACCENT_POSITIVE)
        c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[ms_row].height = 22

    # Legend
    legend_row = ms_row + 2
    ws.cell(row=legend_row, column=2, value="LEGEND:").font = Font(name=FONT_NAME, size=10, bold=HEADER_BOLD, color=PRIMARY)
    ws.cell(row=legend_row, column=2).alignment = Alignment(horizontal='left', vertical='center')

    legend_items = [
        ("Initiate",  ACCENT_WARNING),
        ("Discover",  "6C5B7B"),
        ("Design",    PRIMARY),
        ("Build",     ACCENT_POSITIVE),
        ("Validate",  "D4820A"),
        ("Deploy",    ACCENT_NEGATIVE),
        ("Milestone", ACCENT_POSITIVE),
    ]
    for i, (label, color) in enumerate(legend_items):
        col = 3 + i * 4
        ws.cell(row=legend_row, column=col, value="").fill = PatternFill('solid', fgColor=color)
        c = ws.cell(row=legend_row, column=col + 1, value=label)
        c.font = Font(name=FONT_NAME, size=10, color=NEUTRAL_900)
        c.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[legend_row].height = 22

    ws.sheet_view.zoomScale = 80
    ws.freeze_panes = "G6"


# ============================================================
# Sheet 7 — Payment Milestones
# ============================================================

def build_payment_milestones(wb, totals):
    ws = wb.create_sheet("7. Payment Milestones")
    ws.sheet_view.showGridLines = False

    ws.column_dimensions['A'].width = 3
    widths = {'B': 8, 'C': 40, 'D': 14, 'E': 12, 'F': 14, 'G': 14, 'H': 12, 'I': 36}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    setup_sheet(ws, title="Payment Milestones  —  Billing Schedule (M1–M8)", last_col=9)

    ws.merge_cells('B3:I3')
    sub = ws['B3']
    sub.value = "Milestone-based billing tied to sprint acceptance  |  Net-15 from acceptance  |  USD"
    sub.font = Font(name=FONT_NAME, size=10, color=NEUTRAL_600, italic=True)
    sub.alignment = Alignment(horizontal='left', vertical='center')

    headers = ["ID", "Milestone", "Trigger Sprint", "% of Total", "Invoice Date", "Amount (USD)", "Status", "Acceptance Criteria"]
    header_row = 5
    for i, h in enumerate(headers):
        ws.cell(row=header_row, column=2 + i, value=h)
    style_header_row(ws, row_num=header_row, col_start=2, col_end=9)

    grand_total = totals['grand_total']

    data_start = header_row + 1
    for i, ms in enumerate(MILESTONES):
        mid, mname, sprint, pct, due, status = ms
        r = data_start + i
        amount = round(grand_total * pct / 100, 0)
        ws.cell(row=r, column=2, value=mid).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=r, column=3, value=mname)
        ws.cell(row=r, column=4, value=sprint).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=r, column=5, value=pct / 100).number_format = '0.0%'
        ws.cell(row=r, column=5).alignment = align_number()
        ws.cell(row=r, column=6, value=due).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=r, column=7, value=amount).number_format = '$#,##0'
        ws.cell(row=r, column=7).alignment = align_number()
        ws.cell(row=r, column=8, value=status).alignment = Alignment(horizontal='center', vertical='center')

        # Acceptance criteria
        criteria_map = {
            "M1": "Charter signed; RTM v1 approved; environment strategy accepted",
            "M2": "BRD/PRD signed; stakeholder matrix accepted; baseline plan in place",
            "M3": "SRS/FRS/Solution Arch signed; ADR set accepted; design review passed",
            "M4": "Foundation build deployed to staging; CI/CD green; DevSecOps review passed",
            "M5": "Capture & AI draft beta in UAT hands; confidence policy accepted",
            "M6": "Review/sync governance beta live; Salesforce sandbox integration verified",
            "M7": "UAT sign-off; pilot runbook accepted; performance & security tests passed",
            "M8": "Wave 1 live in production; hypercare closed; pilot KPI validation report accepted",
        }
        ws.cell(row=r, column=9, value=criteria_map.get(mid, "")).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        style_data_row(ws, row_num=r, col_start=2, col_end=9, row_index=i)
        ws.row_dimensions[r].height = 32

    # Totals
    total_row = data_start + len(MILESTONES)
    ws.cell(row=total_row, column=2, value="TOTAL")
    ws.cell(row=total_row, column=3, value="All milestones")
    ws.cell(row=total_row, column=5, value=f"=SUM(E{data_start}:E{total_row - 1})").number_format = '0.0%'
    ws.cell(row=total_row, column=7, value=f"=SUM(G{data_start}:G{total_row - 1})").number_format = '$#,##0'
    style_total_row(ws, row_num=total_row, col_start=2, col_end=9)
    ws.cell(row=total_row, column=5).alignment = align_number()
    ws.cell(row=total_row, column=7).alignment = align_number()
    ws.row_dimensions[total_row].height = 28

    # Status color coding
    ws.conditional_formatting.add(
        f'H{data_start}:H{total_row - 1}',
        CellIsRule(operator='equal', formula=['"Planned"'],
                   font=Font(color=NEUTRAL_600),
                   fill=PatternFill('solid', fgColor=NEUTRAL_100))
    )

    # Payment terms section
    terms_row = total_row + 3
    _section_header(ws, terms_row, 2, 9, "PAYMENT TERMS & CONDITIONS")

    terms = [
        ("Advance",          "50% of M1 invoiced on contract execution; balance on M1 acceptance."),
        ("Subsequent Bills", "Net-15 from milestone acceptance. Late payments >30 days incur 1.5%/month finance charge."),
        ("Acceptance Window","10 business days from delivery notification. Silent acceptance applies if no written rejection."),
        ("Change Control",   "Scope changes >5% of remaining budget require CCB approval; billed as change orders."),
        ("Withholding",      "10% of each milestone retained until M8 acceptance; released on Wave 1 go-live."),
        ("Currency",         "All invoices in USD. Bank charges outside vendor country borne by client."),
        ("Tax",              "Withholding tax (if applicable) deducted at source; vendor provides W-8BEN-E equivalent."),
    ]
    for i, (k, v) in enumerate(terms):
        r = terms_row + 1 + i
        ws.cell(row=r, column=2, value=k).font = Font(name=FONT_NAME, size=10, bold=HEADER_BOLD, color=PRIMARY)
        ws.cell(row=r, column=2).alignment = Alignment(horizontal='left', vertical='center')
        ws.cell(row=r, column=2).fill = PatternFill('solid', fgColor=NEUTRAL_100)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=9)
        ws.cell(row=r, column=3, value=v).font = Font(name=FONT_NAME, size=10, color=NEUTRAL_900)
        ws.cell(row=r, column=3).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws.row_dimensions[r].height = 24

    ws.sheet_view.zoomScale = 100


# ============================================================
# Sheet 8 — Risk & Contingency
# ============================================================

def build_risk_contingency(wb, totals):
    ws = wb.create_sheet("8. Risk & Contingency")
    ws.sheet_view.showGridLines = False

    ws.column_dimensions['A'].width = 3
    widths = {'B': 6, 'C': 44, 'D': 12, 'E': 10, 'F': 14, 'G': 50}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    setup_sheet(ws, title="Risk Register & Contingency Plan", last_col=7)

    ws.merge_cells('B3:G3')
    sub = ws['B3']
    sub.value = f"Reserve: ${totals['reserve']:,.0f}  (8% of base budget ${totals['base_total']:,.0f})  |  Drawn against approved risk realisations"
    sub.font = Font(name=FONT_NAME, size=10, color=NEUTRAL_600, italic=True)
    sub.alignment = Alignment(horizontal='left', vertical='center')

    headers = ["ID", "Risk", "Likelihood", "Impact", "Reserve (USD)", "Mitigation"]
    header_row = 5
    for i, h in enumerate(headers):
        ws.cell(row=header_row, column=2 + i, value=h)
    style_header_row(ws, row_num=header_row, col_start=2, col_end=7)

    data_start = header_row + 1
    for i, risk in enumerate(RISKS):
        rid, rname, like, impact, reserve, mit = risk
        r = data_start + i
        ws.cell(row=r, column=2, value=rid).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=r, column=3, value=rname).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws.cell(row=r, column=4, value=like).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=r, column=5, value=impact).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=r, column=6, value=reserve).number_format = '$#,##0'
        ws.cell(row=r, column=6).alignment = align_number()
        ws.cell(row=r, column=7, value=mit).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        style_data_row(ws, row_num=r, col_start=2, col_end=7, row_index=i)
        ws.row_dimensions[r].height = 42

    # Totals
    total_row = data_start + len(RISKS)
    ws.cell(row=total_row, column=2, value="TOTAL")
    ws.cell(row=total_row, column=3, value="Allocated reserve against R1–R7")
    ws.cell(row=total_row, column=6, value=f"=SUM(F{data_start}:F{total_row - 1})").number_format = '$#,##0'
    style_total_row(ws, row_num=total_row, col_start=2, col_end=7)
    ws.cell(row=total_row, column=6).alignment = align_number()
    ws.row_dimensions[total_row].height = 28

    # Conditional formatting on likelihood
    ws.conditional_formatting.add(
        f'D{data_start}:D{total_row - 1}',
        CellIsRule(operator='equal', formula=['"High"'],
                   font=Font(color=ACCENT_NEGATIVE, bold=HEADER_BOLD),
                   fill=PatternFill('solid', fgColor='FDEDEC'))
    )
    ws.conditional_formatting.add(
        f'D{data_start}:D{total_row - 1}',
        CellIsRule(operator='equal', formula=['"Medium"'],
                   font=Font(color=ACCENT_WARNING, bold=HEADER_BOLD),
                   fill=PatternFill('solid', fgColor='FEF9E7'))
    )
    ws.conditional_formatting.add(
        f'E{data_start}:E{total_row - 1}',
        CellIsRule(operator='equal', formula=['"High"'],
                   font=Font(color=ACCENT_NEGATIVE, bold=HEADER_BOLD),
                   fill=PatternFill('solid', fgColor='FDEDEC'))
    )
    ws.conditional_formatting.add(
        f'E{data_start}:E{total_row - 1}',
        CellIsRule(operator='equal', formula=['"Medium"'],
                   font=Font(color=ACCENT_WARNING, bold=HEADER_BOLD),
                   fill=PatternFill('solid', fgColor='FEF9E7'))
    )

    # Reserve reconciliation
    recon_row = total_row + 3
    _section_header(ws, recon_row, 2, 7, "RESERVE RECONCILIATION")

    rh_row = recon_row + 1
    rh = ["Item", "", "", "", "USD", "Notes"]
    for i, h in enumerate(rh):
        ws.cell(row=rh_row, column=2 + i, value=h)
    style_header_row(ws, row_num=rh_row, col_start=2, col_end=7)
    ws.merge_cells(start_row=rh_row, start_column=2, end_row=rh_row, end_column=5)
    ws.cell(row=rh_row, column=2).value = "Item"

    recon_data = [
        ("Allocated against R1–R7",   sum(r[4] for r in RISKS),        "Sum of reserve column above"),
        ("Unallocated buffer",        totals['reserve'] - sum(r[4] for r in RISKS), "Discretionary; sponsor approval to draw"),
        ("Total Risk Reserve",        totals['reserve'],               "8% of base budget"),
    ]
    for i, (label, usd, notes) in enumerate(recon_data):
        r = rh_row + 1 + i
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
        ws.cell(row=r, column=2, value=label)
        ws.cell(row=r, column=6, value=usd).number_format = '$#,##0'
        ws.cell(row=r, column=6).alignment = align_number()
        ws.cell(row=r, column=7, value=notes).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        is_total = "Total" in label
        if is_total:
            style_total_row(ws, row_num=r, col_start=2, col_end=7)
        else:
            style_data_row(ws, row_num=r, col_start=2, col_end=7, row_index=i)
        ws.row_dimensions[r].height = 24

    # Scenario analysis
    sc_row = rh_row + 1 + len(recon_data) + 3
    _section_header(ws, sc_row, 2, 7, "SCENARIO ANALYSIS  —  Total Cost Outturn")

    sc_headers = ["Scenario", "Description", "", "", "USD", "Δ vs Base"]
    sh_row = sc_row + 1
    for i, h in enumerate(sc_headers):
        ws.cell(row=sh_row, column=2 + i, value=h)
    style_header_row(ws, row_num=sh_row, col_start=2, col_end=7)
    ws.merge_cells(start_row=sh_row, start_column=2, end_row=sh_row, end_column=3)
    ws.cell(row=sh_row, column=2).value = "Scenario"
    ws.merge_cells(start_row=sh_row, start_column=4, end_row=sh_row, end_column=5)
    ws.cell(row=sh_row, column=4).value = "Description"

    scenarios = [
        ("Best Case",   "Risks R1–R7 do not materialise; reserve fully released",                totals['base_total'],                                0),
        ("Base Case",   "Reserve partially drawn (~50%) against R1, R4, R6",                      totals['base_total'] + totals['reserve'] * 0.5,      totals['reserve'] * 0.5),
        ("Worst Case",  "All reserve drawn; +5% scope change order for additional country",       totals['base_total'] + totals['reserve'] + totals['base_total'] * 0.05, totals['reserve'] + totals['base_total'] * 0.05),
    ]
    for i, (sc, desc, usd, delta) in enumerate(scenarios):
        r = sh_row + 1 + i
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        ws.cell(row=r, column=2, value=sc).alignment = Alignment(horizontal='left', vertical='center')
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=5)
        ws.cell(row=r, column=4, value=desc).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws.cell(row=r, column=6, value=usd).number_format = '$#,##0'
        ws.cell(row=r, column=6).alignment = align_number()
        ws.cell(row=r, column=7, value=delta).number_format = '+$#,##0;-$#,##0;$0'
        ws.cell(row=r, column=7).alignment = align_number()
        style_data_row(ws, row_num=r, col_start=2, col_end=7, row_index=i)
        ws.row_dimensions[r].height = 28

    ws.sheet_view.zoomScale = 100


# ============================================================
# Sheet 9 — Assumptions
# ============================================================

def build_assumptions(wb):
    ws = wb.create_sheet("9. Assumptions")
    ws.sheet_view.showGridLines = False

    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 90

    setup_sheet(ws, title="Pricing Basis, Assumptions & Exclusions", last_col=3)

    ws.merge_cells('B3:C3')
    sub = ws['B3']
    sub.value = "Read in conjunction with the master services agreement and statement of work."
    sub.font = Font(name=FONT_NAME, size=10, color=NEUTRAL_600, italic=True)
    sub.alignment = Alignment(horizontal='left', vertical='center')

    header_row = 5
    ws.cell(row=header_row, column=2, value="Topic")
    ws.cell(row=header_row, column=3, value="Detail")
    style_header_row(ws, row_num=header_row, col_start=2, col_end=3)

    for i, (topic, detail) in enumerate(ASSUMPTIONS):
        r = header_row + 1 + i
        ws.cell(row=r, column=2, value=topic)
        ws.cell(row=r, column=3, value=detail)
        style_data_row(ws, row_num=r, col_start=2, col_end=3, row_index=i)
        ws.cell(row=r, column=2).font = Font(name=FONT_NAME, size=10, bold=HEADER_BOLD, color=PRIMARY)
        ws.cell(row=r, column=2).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        ws.cell(row=r, column=3).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        ws.cell(row=r, column=3).font = Font(name=FONT_NAME, size=10, color=NEUTRAL_900)
        ws.row_dimensions[r].height = 38

    # Sign-off block
    sign_row = header_row + 1 + len(ASSUMPTIONS) + 3
    _section_header(ws, sign_row, 2, 3, "ACCEPTANCE & SIGN-OFF")

    sign_data = [
        ("Prepared by",       "Z.ai Delivery — Aceli LAT Pod",                 "2026-06-18"),
        ("Reviewed by",       "________________________________________",       "____________"),
        ("Approved by (Client)", "________________________________________",   "____________"),
        ("Approved by (Vendor)","________________________________________",   "____________"),
    ]
    for i, (role, name, date) in enumerate(sign_data):
        r = sign_row + 1 + i
        ws.cell(row=r, column=2, value=role).font = Font(name=FONT_NAME, size=10, bold=HEADER_BOLD, color=PRIMARY)
        ws.cell(row=r, column=2).alignment = Alignment(horizontal='left', vertical='center')
        ws.cell(row=r, column=2).fill = PatternFill('solid', fgColor=NEUTRAL_100)
        ws.cell(row=r, column=3, value=f"{name}                                  Date: {date}").font = Font(name=FONT_NAME, size=10, color=NEUTRAL_900)
        ws.cell(row=r, column=3).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws.row_dimensions[r].height = 28

    ws.sheet_view.zoomScale = 100


# ============================================================
# Totals pre-computation
# ============================================================

def compute_totals():
    personnel_total = sum(s[6] for s in SPRINTS)
    infra_total = sum(s[7] for s in SPRINTS)
    tools_total = sum(s[8] for s in SPRINTS)
    third_party_total = sum(s[9] for s in SPRINTS)
    travel_total = sum(s[10] for s in SPRINTS)

    # Personnel split by category
    pct_eng = 52.0 / 87.0   # 52 of 87 personnel %
    pct_design = 14.0 / 87.0
    pct_pm = 11.0 / 87.0
    pct_qa = 10.0 / 87.0

    eng_personnel = round(personnel_total * pct_eng, 0)
    design_personnel = round(personnel_total * pct_design, 0)
    pm_personnel = round(personnel_total * pct_pm, 0)
    qa_personnel = round(personnel_total * pct_qa, 0)

    direct_total = infra_total + tools_total + third_party_total + travel_total
    base_total = personnel_total + direct_total
    reserve = round(base_total * 0.08, 0)
    grand_total = base_total + reserve

    return {
        'personnel_total': personnel_total,
        'eng_personnel': eng_personnel,
        'design_personnel': design_personnel,
        'pm_personnel': pm_personnel,
        'qa_personnel': qa_personnel,
        'infra_total': infra_total,
        'tools_total': tools_total,
        'third_party_total': third_party_total,
        'travel_total': travel_total,
        'direct_total': direct_total,
        'base_total': base_total,
        'reserve': reserve,
        'grand_total': grand_total,
    }


# ============================================================
# MAIN
# ============================================================

def main():
    print("Computing totals…")
    totals = compute_totals()
    print(f"  Base budget:   ${totals['base_total']:>12,.0f}")
    print(f"  Risk reserve:  ${totals['reserve']:>12,.0f}  (8%)")
    print(f"  Grand total:   ${totals['grand_total']:>12,.0f}")
    print()

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    print("Building Cover…")
    build_cover(wb)
    print("Building Executive Summary…")
    build_executive_summary(wb, totals)
    print("Building Sprint Budget…")
    build_sprint_budget(wb, totals)
    print("Building Resource Costing…")
    build_resource_costing(wb, totals)
    print("Building Cost by Category…")
    build_cost_by_category(wb, totals)
    print("Building Development Timeline…")
    build_timeline(wb)
    print("Building Payment Milestones…")
    build_payment_milestones(wb, totals)
    print("Building Risk & Contingency…")
    build_risk_contingency(wb, totals)
    print("Building Assumptions…")
    build_assumptions(wb)

    wb.properties.creator = "Z.ai"
    wb.properties.title = "Aceli LAT — Budget, Timeline & Costing Workbook"
    wb.properties.subject = "Master budget for Aceli LAT engagement (Sprint 0–7)"

    out_path = "/home/z/my-project/download/Aceli_LAT_Budget_Timeline_Costing.xlsx"
    wb.save(out_path)
    print(f"\n✓ Saved: {out_path}")
    print(f"  Sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()
